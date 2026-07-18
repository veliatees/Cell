#!/usr/bin/env python3
"""Build the compact human-liver open-data atlas used by the engine.

The source workbooks are downloaded only while this explicit curation command
runs. Checksums are verified before any values are parsed; the repository keeps
the much smaller, reviewable JSON result rather than multi-gigabyte raw atlases.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import statistics
import urllib.request
from collections import Counter, defaultdict
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from zipfile import ZipFile

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - exercised by operators
    raise SystemExit(
        "openpyxl is required for this explicit source-curation command"
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = (
    ROOT / "data/phh_baseline/curated/human_liver_open_atlas.v1.json"
)
DEFAULT_CACHE = Path("/tmp/cell-human-liver-open-atlas-sources")
DATE_VERIFIED = "2026-07-15"


SOURCES: dict[str, dict[str, Any]] = {
    "fabyan2026_liver_3d": {
        "title": "3D reconstruction of human liver tissue at cellular resolution",
        "paper_url": "https://doi.org/10.1126/sciadv.adz2299",
        "artifact_url": (
            "https://zenodo.org/records/18479511/files/"
            "Fabyan_et_al_ScienceAdvances.zip?download=1"
        ),
        "filename": "Fabyan_et_al_ScienceAdvances.zip",
        "expected_size_bytes": 60_997_948,
        "expected_md5": "f32c852d6f5bc87ca2b8e239f6f5b436",
        "license": "CC-BY-4.0",
        "source_role": "human_tissue_architecture_and_central_vein_morphometry",
    },
    "watson2025_human_liver_source_data": {
        "title": "Spatial transcriptomics of healthy and fibrotic human liver at single-cell resolution",
        "paper_url": "https://www.nature.com/articles/s41467-024-55325-4",
        "artifact_url": (
            "https://static-content.springer.com/esm/art%3A10.1038%2F"
            "s41467-024-55325-4/MediaObjects/41467_2024_55325_MOESM11_ESM.xlsx"
        ),
        "filename": "41467_2024_55325_MOESM11_ESM.xlsx",
        "expected_size_bytes": 2_922_301,
        "expected_md5": "a9368e708e6285c436cf82a6d273b439",
        "license": "CC-BY-4.0",
        "source_role": "healthy_human_hepatocyte_2d_morphometry",
    },
    "watson2025_cellphonedb": {
        "title": "Spatial transcriptomics of healthy and fibrotic human liver at single-cell resolution",
        "paper_url": "https://www.nature.com/articles/s41467-024-55325-4",
        "artifact_url": (
            "https://static-content.springer.com/esm/art%3A10.1038%2F"
            "s41467-024-55325-4/MediaObjects/41467_2024_55325_MOESM6_ESM.xlsx"
        ),
        "filename": "41467_2024_55325_MOESM6_ESM.xlsx",
        "expected_size_bytes": 486_918,
        "expected_md5": "cbc0a0fdb50114c9d29fea357d47c44e",
        "license": "CC-BY-4.0",
        "source_role": "cellphonedb_interaction_hypotheses",
    },
    "mallanna2016_phh_surfaceome": {
        "title": (
            "Mapping the Cell-Surface N-Glycoproteome of Human Hepatocytes "
            "Reveals Markers for Selecting a Homogeneous Population of "
            "iPSC-Derived Hepatocytes"
        ),
        "paper_url": "https://pmc.ncbi.nlm.nih.gov/articles/PMC5032032/",
        "artifact_url": (
            "https://ars.els-cdn.com/content/image/"
            "1-s2.0-S2213671116301400-mmc2.xlsx"
        ),
        "filename": "S2213671116301400-mmc2.xlsx",
        "expected_size_bytes": 3_124_245,
        "expected_md5": "5d6ff2cf3fb356f9b335b598da6512e6",
        "license": "CC-BY-4.0",
        "source_role": "primary_human_hepatocyte_surface_nglycoproteome",
    },
    "weiss2026_spatial_proteome": {
        "title": (
            "Single-cell spatial proteomics maps human liver zonation patterns "
            "and their vulnerability to disruption in tissue architecture"
        ),
        "paper_url": "https://www.nature.com/articles/s42255-026-01459-2",
        "artifact_url": (
            "https://static-content.springer.com/esm/art%3A10.1038%2F"
            "s42255-026-01459-2/MediaObjects/42255_2026_1459_MOESM3_ESM.xlsx"
        ),
        "filename": "42255_2026_1459_MOESM3_ESM.xlsx",
        "expected_size_bytes": 638_091,
        "expected_md5": "2da07de01f671a234b5c9ace65fc137b",
        "license": "CC-BY-4.0",
        "source_role": "healthy_human_hepatocyte_spatial_proteomics",
    },
}


def _digest(path: Path, algorithm: str) -> str:
    hasher = hashlib.new(algorithm)
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _fetch_source(source_id: str, cache_dir: Path) -> tuple[Path, dict[str, Any]]:
    source = SOURCES[source_id]
    cache_dir.mkdir(parents=True, exist_ok=True)
    target = cache_dir / source["filename"]
    if not target.exists():
        request = urllib.request.Request(
            source["artifact_url"],
            headers={"User-Agent": "Cell-open-atlas-curator/1.0"},
        )
        with urllib.request.urlopen(request, timeout=180) as response:
            target.write_bytes(response.read())

    actual_size = target.stat().st_size
    actual_md5 = _digest(target, "md5")
    if actual_size != source["expected_size_bytes"]:
        raise ValueError(
            f"{source_id} size mismatch: {actual_size} != "
            f"{source['expected_size_bytes']}"
        )
    if actual_md5 != source["expected_md5"]:
        raise ValueError(
            f"{source_id} MD5 mismatch: {actual_md5} != "
            f"{source['expected_md5']}"
        )

    artifact = {
        "id": source_id,
        "title": source["title"],
        "paper_url": source["paper_url"],
        "artifact_url": source["artifact_url"],
        "filename": source["filename"],
        "size_bytes": actual_size,
        "md5": actual_md5,
        "sha256": _digest(target, "sha256"),
        "license": source["license"],
        "source_role": source["source_role"],
        "date_verified": DATE_VERIFIED,
    }
    return target, artifact


def _percentile(values: list[float], fraction: float) -> float:
    if not values:
        raise ValueError("percentile requires at least one value")
    ordered = sorted(values)
    index = (len(ordered) - 1) * fraction
    lower = math.floor(index)
    upper = math.ceil(index)
    if lower == upper:
        return ordered[lower]
    weight = index - lower
    return ordered[lower] * (1.0 - weight) + ordered[upper] * weight


def _distribution(values: Iterable[float]) -> dict[str, float | int]:
    cleaned = [float(value) for value in values]
    if not cleaned:
        raise ValueError("distribution requires at least one finite value")
    if not all(math.isfinite(value) for value in cleaned):
        raise ValueError("distribution contains a non-finite value")
    return {
        "count": len(cleaned),
        "mean": statistics.fmean(cleaned),
        "sample_sd": statistics.stdev(cleaned) if len(cleaned) > 1 else 0.0,
        "p05": _percentile(cleaned, 0.05),
        "p25": _percentile(cleaned, 0.25),
        "median": _percentile(cleaned, 0.5),
        "p75": _percentile(cleaned, 0.75),
        "p95": _percentile(cleaned, 0.95),
        "minimum": min(cleaned),
        "maximum": max(cleaned),
    }


def _source_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    match = re.fullmatch(r"([+-]?[0-9.]+)\*10([+-]?[0-9]+)", text)
    if match:
        return float(match.group(1)) * 10 ** int(match.group(2))
    return float(text)


def _curate_tissue_architecture(path: Path) -> dict[str, Any]:
    member = "Fabyan_et_al_ScienceAdvances/Figure4/4C.xlsx"
    with ZipFile(path) as archive:
        workbook = load_workbook(
            BytesIO(archive.read(member)), read_only=True, data_only=True
        )
    sheet = workbook.active
    control: list[dict[str, float | int]] = []
    cirrhotic: list[dict[str, float | int]] = []
    column_sets = {
        "control": (1, 4, 7, 10, 13),
        "cirrhotic": (2, 5, 8, 11, 14),
    }
    for row_index in range(3, sheet.max_row + 1):
        for cohort, columns in column_sets.items():
            values = [_source_number(sheet.cell(row_index, column).value) for column in columns]
            if values[0] is None:
                continue
            record = {
                "venule_count": int(values[0]),
                "total_volume_um3": float(values[1] or 0.0),
                "total_length_um": float(values[2] or 0.0),
                "branching_node_count": int(values[3] or 0),
                "ending_count": int(values[4] or 0),
            }
            (control if cohort == "control" else cirrhotic).append(record)

    return {
        "source_id": "fabyan2026_liver_3d",
        "reconstructed_tissue_extent_um": [4000.0, 4000.0, 500.0],
        "healthy_lobule_polygonal_radius_um": {
            "count": 17,
            "median": 595.0,
            "minimum": 428.0,
            "maximum": 717.0,
            "value_status": "source_reported_3d_human_tissue",
        },
        "independent_2d_histology_lobule_radius_um": {
            "measurement_count": 52,
            "sample_count": 6,
            "mean": 592.0,
            "sample_sd": 87.0,
            "value_status": "source_reported_nonfibrotic_human_histology",
        },
        "central_vein_network_source_rows": {
            "healthy_control": control,
            "cirrhotic_context_only": cirrhotic,
        },
        "healthy_initialization_may_use_cirrhotic_rows": False,
        "limitations": [
            "Lobule-scale values constrain future tissue geometry, not single-cell dimensions.",
            "The central-vein workbook supplies sample-level networks, not a full sinusoid mesh.",
            "Cirrhotic rows are retained as separate validation context and cannot initialize a healthy scene.",
        ],
    }


def _curate_hepatocyte_morphometry(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Figure 4"]
    areas_by_cluster: dict[str, list[float]] = defaultdict(list)
    areas_by_nuclei: dict[str, list[float]] = defaultdict(list)
    detected_nuclei = Counter()
    records = 0
    for row in sheet.iter_rows(min_row=2, min_col=6, max_col=11, values_only=True):
        cell_id, cluster, nuclei, area, _counts, _density = row
        if not cell_id or cluster not in {"Hep_1", "Hep_2", "Hep_3"}:
            continue
        if not isinstance(area, (int, float)) or not (
            isinstance(nuclei, (int, float)) or nuclei == ">=4"
        ):
            raise ValueError(f"invalid morphometry row for {cell_id}")
        records += 1
        area_value = float(area)
        nuclei_key = ">=4" if nuclei == ">=4" else str(int(nuclei))
        areas_by_cluster[str(cluster)].append(area_value)
        areas_by_nuclei[nuclei_key].append(area_value)
        detected_nuclei[nuclei_key] += 1

    all_areas = [value for values in areas_by_cluster.values() for value in values]
    if records != 56_055:
        raise ValueError(f"unexpected healthy hepatocyte count: {records}")
    if sum(detected_nuclei.values()) != records:
        raise ValueError("nucleus-detection categories do not cover all hepatocytes")

    return {
        "source_id": "watson2025_human_liver_source_data",
        "assay": "2d_in_situ_segmentation_of_healthy_human_liver_imaging",
        "cell_count": records,
        "cluster_semantics": {
            "Hep_1": "source-defined healthy hepatocyte cluster mapped to zone 1",
            "Hep_2": "source-defined healthy hepatocyte cluster mapped between zones 1 and 3",
            "Hep_3": "source-defined healthy hepatocyte cluster mapped to zone 3",
        },
        "source_reported_cluster_to_zone": {
            "Hep_1": "periportal",
            "Hep_2": "midlobular",
            "Hep_3": "pericentral",
        },
        "cluster_zone_mapping_status": "source_reported_for_this_healthy_merfish_cohort",
        "segmented_area_um2": {
            "all": _distribution(all_areas),
            "by_cluster": {
                cluster: _distribution(values)
                for cluster, values in sorted(areas_by_cluster.items())
            },
            "by_detected_nuclei": {
                key: _distribution(areas_by_nuclei[key])
                for key in ("0", "1", "2", "3", ">=4")
            },
        },
        "detected_nuclei_count": {
            "counts": {key: detected_nuclei[key] for key in ("0", "1", "2", "3", ">=4")},
            "fractions": {
                key: detected_nuclei[key] / records
                for key in ("0", "1", "2", "3", ">=4")
            },
            "zero_is_segmentation_nonassignment_not_biological_anucleation": True,
        },
        "renderer_distribution_available": True,
        "may_replace_3d_cell_geometry": False,
        "limitations": [
            "Segmented area is a two-dimensional in-situ observation, not cell volume or surface area.",
            "The source maps Hep_1/Hep_2/Hep_3 to zones 1/2/3; that cohort-specific mapping is not a universal donor classifier.",
            "A zero detected-nucleus count does not demonstrate an anucleate hepatocyte.",
        ],
    }


def _curate_surfaceome(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["2) Protein Names_Annotations"]
    proteins: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=3, max_col=6, values_only=True):
        index, accession, gene, protein_name, type_annotation, cd_molecule = row
        if not isinstance(index, (int, float)):
            continue
        proteins.append(
            {
                "source_index": int(index),
                "accession": str(accession),
                "gene": str(gene),
                "protein_name": str(protein_name),
                "source_type_annotation": str(type_annotation or ""),
                "cd_molecule": str(cd_molecule) if cd_molecule else None,
                "observed_by_cell_surface_capture": True,
                "surface_density_per_um2": None,
                "membrane_domain": None,
                "orientation": None,
            }
        )
    if len(proteins) != 300 or len({item["gene"] for item in proteins}) != 300:
        raise ValueError("surfaceome source must contain 300 unique gene identities")

    reported_cd_count = sum(item["cd_molecule"] is not None for item in proteins)
    reported_tm_count = sum(
        "TM" in item["source_type_annotation"].upper() for item in proteins
    )
    if reported_cd_count != 66 or reported_tm_count != 228:
        raise ValueError(
            "surfaceome annotations changed: "
            f"CD={reported_cd_count}, transmembrane={reported_tm_count}"
        )
    genes = {item["gene"] for item in proteins}
    relevant = (
        "SLC10A1",
        "EGFR",
        "INSR",
        "MET",
        "IL6ST",
        "ABCB11",
        "ABCC2",
        "ABCB1",
        "CDH1",
        "GJB1",
        "GCGR",
        "IL6R",
        "FZD6",
        "LRP5",
        "LRP6",
    )
    return {
        "source_id": "mallanna2016_phh_surfaceome",
        "assay": "cell_surface_capture_of_n_glycosylated_proteins_in_primary_human_hepatocytes",
        "observed_protein_count": len(proteins),
        "reported_cd_molecule_count": reported_cd_count,
        "reported_transmembrane_count": reported_tm_count,
        "proteins": proteins,
        "pathway_relevant_gene_observation": {
            gene: "observed" if gene in genes else "not_detected_in_this_assay"
            for gene in relevant
        },
        "absence_is_proof_of_biological_absence": False,
        "density_available": False,
        "membrane_domain_available": False,
        "orientation_available": False,
        "limitations": [
            "Cell-surface capture establishes assay-level identity, not molecules per square micrometre.",
            "A protein not detected by this assay cannot be declared absent from hepatocytes.",
            "The table does not assign sinusoidal, lateral or canalicular membrane domains.",
        ],
    }


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"yes", "true", "1"}


def _curate_spatial_proteome(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Table 3"]
    records: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=6, max_col=27, values_only=True):
        protein = row[0]
        if not protein:
            continue
        bins = [float(value) for value in row[1:21]]
        coefficient = float(row[22])
        zonated = _as_bool(row[25])
        strong_zonated = _as_bool(row[26])
        enriched_region = (
            "periportal" if coefficient > 0 else "pericentral" if coefficient < 0 else "flat"
        )
        records.append(
            {
                "protein": str(protein),
                "binned_expression_percent": bins,
                "binned_expression_sum": float(row[21]),
                "coefficient": coefficient,
                "p_value": float(row[23]),
                "q_value": float(row[24]),
                "zonated": zonated,
                "strong_zonated": strong_zonated,
                "enriched_region": enriched_region,
            }
        )
    if not records:
        raise ValueError("published spatial-proteome table contains no records")
    strong = [record for record in records if record["strong_zonated"]]
    portal = [record for record in strong if record["enriched_region"] == "periportal"]
    central = [record for record in strong if record["enriched_region"] == "pericentral"]
    if len(strong) != 171 or len(portal) != 102 or len(central) != 69:
        raise ValueError(
            "published strong-zonation classification changed: "
            f"all={len(strong)}, portal={len(portal)}, central={len(central)}"
        )
    return {
        "source_id": "weiss2026_spatial_proteome",
        "cohort": "healthy_human_liver_N14",
        "analyzed_hepatocyte_count_after_filtering": 413,
        "porto_central_bin_count": 20,
        "bin_zero_region": "central",
        "bin_nineteen_region": "portal",
        "protein_count": len(records),
        "article_reported_protein_count_at_70pct_completeness": 1_741,
        "supplement_table_record_count": len(records),
        "article_minus_supplement_record_count": 1_741 - len(records),
        "strong_zonated_count": len(strong),
        "strong_periportal_count": len(portal),
        "strong_pericentral_count": len(central),
        "strong_definition": "abs(coefficient) > 1 and q_value < 0.05",
        "records": records,
        "may_scale_metabolic_flux": False,
        "limitations": [
            "Binned expression values are normalized within protein and are not absolute copy numbers.",
            "The article reports 1,741 proteins at >=70% completeness, while the published supplementary table contains 1,736 machine-readable protein rows; the five-record difference is unresolved.",
            "The zonation coefficient is not a reaction-rate or transport multiplier.",
            "The monotonic portal-central analysis does not define a separate midlobular maximum class.",
        ],
    }


def _curate_interactions(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["statistical_analysis_interactio"]
    rows = sheet.iter_rows(values_only=True)
    header = tuple(str(value) for value in next(rows))
    pair_columns = header[13:]
    retained: list[dict[str, Any]] = []
    nonzero_edge_count = 0
    for source_row in rows:
        edges: list[dict[str, Any]] = []
        for pair, value in zip(pair_columns, source_row[13:]):
            score = float(value or 0.0)
            if score <= 0.0:
                continue
            sender, receiver = pair.split("|", maxsplit=1)
            if not (sender.startswith("Hep_") or receiver.startswith("Hep_")):
                continue
            edges.append({"sender": sender, "receiver": receiver, "score": score})
        if not edges:
            continue
        retained.append(
            {
                "id": str(source_row[0]),
                "interacting_pair": str(source_row[1]),
                "partner_a": str(source_row[2]),
                "partner_b": str(source_row[3]),
                "gene_a": str(source_row[4]) if source_row[4] else None,
                "gene_b": str(source_row[5]) if source_row[5] else None,
                "secreted": bool(source_row[6]),
                "receptor_a": bool(source_row[7]),
                "receptor_b": bool(source_row[8]),
                "annotation_strategy": str(source_row[9]),
                "is_integrin": bool(source_row[10]),
                "directionality": str(source_row[11]),
                "classification": str(source_row[12]),
                "hepatocyte_edges": edges,
            }
        )
        nonzero_edge_count += len(edges)

    if len(retained) != 209 or nonzero_edge_count != 1_806:
        raise ValueError(
            "hepatocyte interaction subset changed: "
            f"interactions={len(retained)}, edges={nonzero_edge_count}"
        )
    return {
        "source_id": "watson2025_cellphonedb",
        "method": "CellPhoneDB_source_interaction_score",
        "source_interaction_count": sheet.max_row - 1,
        "retained_hepatocyte_interaction_count": len(retained),
        "nonzero_hepatocyte_edge_count": nonzero_edge_count,
        "cell_types": ["HSC_1", "HSC_2", "Hep_1", "Hep_2", "Hep_3", "LSEC", "Mac_1", "Mac_2"],
        "source_reported_hepatocyte_cluster_to_zone": {
            "Hep_1": "periportal",
            "Hep_2": "midlobular",
            "Hep_3": "pericentral",
        },
        "records": retained,
        "score_range": [0.0, 100.0],
        "score_is_binding_probability": False,
        "score_is_kinetic_rate": False,
        "may_activate_contact_chain": False,
        "limitations": [
            "CellPhoneDB scores nominate potential communication edges; they do not prove physical binding.",
            "The source-reported Hep_1/Hep_2/Hep_3 zone mapping is cohort-specific and cannot be treated as a universal classifier.",
            "Local membrane density, orientation and two-dimensional on/off kinetics remain required for activation.",
        ],
    }


def build_atlas(cache_dir: Path) -> dict[str, Any]:
    paths: dict[str, Path] = {}
    artifacts: list[dict[str, Any]] = []
    for source_id in SOURCES:
        path, artifact = _fetch_source(source_id, cache_dir)
        paths[source_id] = path
        artifacts.append(artifact)

    return {
        "schema_version": "cell.human-liver-open-atlas.v1",
        "status": "curated_primary_human_evidence_with_fail_closed_quantitative_gates",
        "date_verified": DATE_VERIFIED,
        "generated_on": DATE_VERIFIED,
        "source_artifacts": artifacts,
        "tissue_architecture": _curate_tissue_architecture(paths["fabyan2026_liver_3d"]),
        "hepatocyte_morphometry_2d": _curate_hepatocyte_morphometry(
            paths["watson2025_human_liver_source_data"]
        ),
        "surface_nglycoproteome": _curate_surfaceome(
            paths["mallanna2016_phh_surfaceome"]
        ),
        "spatial_proteome": _curate_spatial_proteome(
            paths["weiss2026_spatial_proteome"]
        ),
        "hepatocyte_interaction_hypotheses": _curate_interactions(
            paths["watson2025_cellphonedb"]
        ),
        "integration_gates": {
            "may_sample_2d_renderer_area_distribution": True,
            "may_replace_3d_cell_geometry": False,
            "may_use_surface_protein_identity": True,
            "surface_density_available": False,
            "membrane_domain_available": False,
            "surface_orientation_available": False,
            "may_display_spatial_protein_gradient": True,
            "may_scale_flux_from_spatial_proteome": False,
            "may_rank_interaction_hypotheses": True,
            "may_activate_interaction_from_score": False,
            "binding_kinetics_available": False,
        },
        "global_limitations": [
            "This bundle joins compatible evidence layers without pretending that they came from one donor or one assay.",
            "Identity, normalized zonation and interaction scores are not interchangeable with absolute abundance or kinetics.",
            "No missing density, orientation, copy number, three-dimensional morphology or rate constant is imputed.",
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cache-dir", type=Path, default=DEFAULT_CACHE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    arguments = parser.parse_args()
    atlas = build_atlas(arguments.cache_dir)
    arguments.output.parent.mkdir(parents=True, exist_ok=True)
    arguments.output.write_text(
        json.dumps(atlas, indent=2, ensure_ascii=True) + "\n", encoding="utf-8"
    )
    print(
        f"wrote {arguments.output} "
        f"({atlas['hepatocyte_morphometry_2d']['cell_count']} hepatocytes, "
        f"{atlas['surface_nglycoproteome']['observed_protein_count']} surface proteins, "
        f"{atlas['spatial_proteome']['protein_count']} spatial proteins, "
        f"{atlas['hepatocyte_interaction_hypotheses']['retained_hepatocyte_interaction_count']} interactions)"
    )


if __name__ == "__main__":
    main()
