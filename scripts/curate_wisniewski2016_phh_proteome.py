#!/usr/bin/env python3
"""Curate the seven-donor absolute PHH proteome from the official supplements.

The raw Elsevier workbooks are checksum-locked and are not redistributed. The
repository stores only factual measurements needed by the engine. Source zeros
and blanks are retained as non-quantified ``null`` values; this command never
imputes a protein abundance and never collapses distinct MaxQuant protein groups.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import statistics
import urllib.request
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - operator environment
    raise SystemExit("openpyxl is required for PHH proteome curation") from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = (
    ROOT
    / "data/phh_baseline/curated"
    / "wisniewski2016_donor_proteome_atlas.v1.json"
)
DEFAULT_CACHE = Path("/tmp/cell-wisniewski2016-proteome")
DATE_VERIFIED = "2026-07-16"
DONOR_IDS = tuple("ABCDEFG")

SOURCE_FILES: dict[str, dict[str, Any]] = {
    "supplemental_table_1": {
        "title": "Supplementary Table 1: hepatocyte donor characteristics",
        "url": "https://ars.els-cdn.com/content/image/1-s2.0-S1874391916300197-mmc1.xlsx",
        "filename": "1-s2.0-S1874391916300197-mmc1.xlsx",
        "expected_size_bytes": 10_366,
        "expected_sha256": "9bbc90323a184d8224388b927d720343e6222b5f58b4eba8355a64e4b918f17a",
        "source_role": "donor_context",
    },
    "supplemental_table_2": {
        "title": "Supplementary Table 2: quantitative hepatocyte proteome",
        "url": "https://ars.els-cdn.com/content/image/1-s2.0-S1874391916300197-mmc2.xlsx",
        "filename": "1-s2.0-S1874391916300197-mmc2.xlsx",
        "expected_size_bytes": 15_457_204,
        "expected_sha256": "f84b9c2a4af4cac3ba6394907e50786485c789ab5ed6421de76bf0d52ebb46d0",
        "source_role": "donor_resolved_absolute_protein_abundance",
    },
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _verify(path: Path, source_id: str) -> dict[str, Any]:
    source = SOURCE_FILES[source_id]
    actual_size = path.stat().st_size
    actual_sha256 = _sha256(path)
    if actual_size != source["expected_size_bytes"]:
        raise ValueError(
            f"{source_id} size mismatch: {actual_size} != "
            f"{source['expected_size_bytes']}"
        )
    if actual_sha256 != source["expected_sha256"]:
        raise ValueError(
            f"{source_id} SHA-256 mismatch: {actual_sha256} != "
            f"{source['expected_sha256']}"
        )
    return {
        "id": source_id,
        "title": source["title"],
        "url": source["url"],
        "filename": source["filename"],
        "size_bytes": actual_size,
        "sha256": actual_sha256,
        "source_role": source["source_role"],
        "reuse_boundary": (
            "factual measurements curated from the official supplement; "
            "raw workbook not redistributed"
        ),
        "date_verified": DATE_VERIFIED,
    }


def _fetch(source_id: str, cache_dir: Path) -> Path:
    source = SOURCE_FILES[source_id]
    cache_dir.mkdir(parents=True, exist_ok=True)
    target = cache_dir / source["filename"]
    if not target.exists():
        request = urllib.request.Request(
            source["url"], headers={"User-Agent": "Cell-PHH-proteome-curator/1.0"}
        )
        with urllib.request.urlopen(request, timeout=180) as response:
            target.write_bytes(response.read())
    _verify(target, source_id)
    return target


def _parts(value: object) -> list[str]:
    return [part for part in str(value or "").split(";") if part]


def _positive_or_none(value: object) -> float | None:
    if not isinstance(value, (int, float)):
        return None
    number = float(value)
    return number if number > 0.0 else None


def _mean(values: list[float]) -> float:
    if not values:
        raise ValueError("mean requires at least one value")
    return statistics.fmean(values)


def _curate_donors(table_1: Path) -> list[dict[str, object]]:
    workbook = load_workbook(table_1, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if rows[0] != ("Hepatocyte Sample", "Age", "Gender", "Diagnosis"):
        raise ValueError("supplemental donor-table headers changed")
    donors = [
        {
            "id": str(donor_id),
            "age_years": int(age),
            "sex_as_reported": str(sex),
            "diagnosis_as_reported": str(diagnosis),
            "tissue_context": "histologically_normal_surgical_resection_tissue",
        }
        for donor_id, age, sex, diagnosis in rows[1:]
    ]
    if tuple(item["id"] for item in donors) != DONOR_IDS:
        raise ValueError("supplemental donor identifiers changed")
    return donors


def _total_protein_per_nucleus(
    workbook: Any,
) -> dict[str, dict[str, object]]:
    sheet = workbook["Histone ruler"]
    headers = tuple(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    mass_row = next(
        row
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if str(row[0] or "").strip() == "Total protein / nucleus (pg)"
    )
    values: dict[str, list[float]] = {donor_id: [] for donor_id in DONOR_IDS}
    for header, raw in zip(headers[2:], mass_row[2:]):
        match = re.fullmatch(r"Total protein ([A-G]).*", str(header or ""))
        if match and isinstance(raw, (int, float)) and float(raw) > 0.0:
            values[match.group(1)].append(float(raw))
    if any(not donor_values for donor_values in values.values()):
        raise ValueError("histone-ruler total-protein rows are incomplete")
    return {
        donor_id: {
            "replicate_values_pg_per_nucleus": donor_values,
            "replicate_count": len(donor_values),
            "mean_pg_per_nucleus": _mean(donor_values),
            "minimum_pg_per_nucleus": min(donor_values),
            "maximum_pg_per_nucleus": max(donor_values),
        }
        for donor_id, donor_values in values.items()
    }


def _normalized_header(value: object) -> str:
    return " ".join(str(value or "").split())


def _curate_protein_groups(
    workbook: Any,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    sheet = workbook["Protein groups"]
    headers = tuple(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_index = {_normalized_header(value): index for index, value in enumerate(headers)}
    required = {
        "Protein IDs",
        "Majority protein IDs",
        "Protein names",
        "Gene names",
        "Peptides",
        "Razor + unique peptides",
        "Unique peptides",
        "Mol. weight [kDa]",
        "PEP",
    }
    required.update(f"Average concentration {donor_id}" for donor_id in DONOR_IDS)
    required.update(f"Averagecopy number {donor_id}" for donor_id in DONOR_IDS)
    if not required <= set(header_index):
        raise ValueError(
            f"supplemental proteome headers changed: {sorted(required - set(header_index))}"
        )

    records: list[dict[str, object]] = []
    source_counts = Counter()
    coverage = Counter()
    donor_group_counts = Counter()
    donor_copy_sums = Counter()
    group_ids: set[str] = set()

    for source_row, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        source_counts["source_rows"] += 1
        source_protein_ids = _parts(row[header_index["Protein IDs"]])
        target_ids = [
            item
            for item in source_protein_ids
            if not item.startswith(("CON__", "REV__"))
        ]
        if not target_ids:
            category = (
                "reverse_only_rows"
                if source_protein_ids
                and all(item.startswith("REV__") for item in source_protein_ids)
                else "contaminant_only_rows"
            )
            source_counts[category] += 1
            continue
        source_counts["target_rows"] += 1

        majority_ids = [
            item
            for item in _parts(row[header_index["Majority protein IDs"]])
            if not item.startswith(("CON__", "REV__"))
        ]
        donor_values: dict[str, dict[str, float | None]] = {}
        detected = 0
        for donor_id in DONOR_IDS:
            concentration = _positive_or_none(
                row[header_index[f"Average concentration {donor_id}"]]
            )
            copies = _positive_or_none(
                row[header_index[f"Averagecopy number {donor_id}"]]
            )
            if (concentration is None) != (copies is None):
                raise ValueError(
                    f"concentration/copy missingness mismatch at source row {source_row}"
                )
            donor_values[donor_id] = {
                "concentration_pmol_per_mg_total_protein": concentration,
                "copies_per_nucleus": copies,
            }
            if copies is not None:
                detected += 1
                donor_group_counts[donor_id] += 1
                donor_copy_sums[donor_id] += copies
        if detected == 0:
            source_counts["target_rows_without_positive_phh_value"] += 1
            continue

        source_counts["quantified_target_rows"] += 1
        coverage[str(detected)] += 1
        group_id = (majority_ids or target_ids)[0]
        if group_id in group_ids:
            group_id = f"{group_id}@row{source_row}"
        group_ids.add(group_id)
        records.append(
            {
                "group_id": group_id,
                "source_row": source_row,
                "protein_ids": target_ids,
                "majority_protein_ids": majority_ids,
                "protein_names": _parts(row[header_index["Protein names"]]),
                "gene_names": _parts(row[header_index["Gene names"]]),
                "identification": {
                    "peptides": int(row[header_index["Peptides"]]),
                    "razor_plus_unique_peptides": int(
                        row[header_index["Razor + unique peptides"]]
                    ),
                    "unique_peptides": int(row[header_index["Unique peptides"]]),
                    "molecular_weight_kda": float(
                        row[header_index["Mol. weight [kDa]"]]
                    ),
                    "posterior_error_probability": float(row[header_index["PEP"]]),
                },
                "detected_donor_count": detected,
                "had_contaminant_member_in_source_group": any(
                    item.startswith("CON__") for item in source_protein_ids
                ),
                "donor_values": donor_values,
            }
        )

    expected_counts = {
        "source_rows": 9_565,
        "target_rows": 9_386,
        "contaminant_only_rows": 179,
        "quantified_target_rows": 8_689,
        "target_rows_without_positive_phh_value": 697,
    }
    if any(source_counts[key] != value for key, value in expected_counts.items()):
        raise ValueError(f"protein-group source audit changed: {dict(source_counts)}")
    if source_counts["reverse_only_rows"] != 0:
        raise ValueError("unexpected reverse-only protein groups appeared")

    return records, {
        **expected_counts,
        "article_reported_whole_cell_lysate_protein_count": 8_705,
        "article_reported_combined_dataset_protein_count": 9_400,
        "reverse_only_rows": 0,
        "detected_donor_coverage_histogram": {
            str(count): coverage[str(count)] for count in range(1, 8)
        },
        "donor_quantified_group_count": {
            donor_id: donor_group_counts[donor_id] for donor_id in DONOR_IDS
        },
        "donor_sum_of_quantified_target_group_copies_per_nucleus": {
            donor_id: donor_copy_sums[donor_id] for donor_id in DONOR_IDS
        },
    }


def curate(table_1: Path, table_2: Path) -> dict[str, object]:
    artifacts = [
        _verify(table_1, "supplemental_table_1"),
        _verify(table_2, "supplemental_table_2"),
    ]
    donors = _curate_donors(table_1)
    workbook = load_workbook(table_2, read_only=True, data_only=True)
    mass_by_donor = _total_protein_per_nucleus(workbook)
    records, source_audit = _curate_protein_groups(workbook)
    for donor in donors:
        donor_id = str(donor["id"])
        donor["total_protein_measurement"] = mass_by_donor[donor_id]
        donor["quantified_target_group_count"] = source_audit[
            "donor_quantified_group_count"
        ][donor_id]
        donor["sum_of_quantified_target_group_copies_per_nucleus"] = source_audit[
            "donor_sum_of_quantified_target_group_copies_per_nucleus"
        ][donor_id]

    donor_mean_mass = _mean(
        [
            float(donor["total_protein_measurement"]["mean_pg_per_nucleus"])
            for donor in donors
        ]
    )
    donor_mean_copy_sum = _mean(
        [float(donor["sum_of_quantified_target_group_copies_per_nucleus"]) for donor in donors]
    )
    return {
        "schema_version": "cell.phh-absolute-proteome-atlas.v1",
        "version": "phh_absolute_proteome_atlas_v1",
        "status": "donor_resolved_static_absolute_abundance_reference",
        "date_verified": DATE_VERIFIED,
        "source": {
            "paper_title": (
                "In-depth quantitative analysis and comparison of the human "
                "hepatocyte and hepatoma cell line HepG2 proteomes"
            ),
            "doi": "10.1016/j.jprot.2016.01.016",
            "pubmed_id": "26825538",
            "proteomics_repository": {
                "accession": "MSV000079562",
                "proteomexchange_accession": "PXD001874",
                "license": "CC0-1.0",
                "url": "https://massive.ucsd.edu/ProteoSAFe/dataset.jsp?task=2ed487a661bf401caae8285acc1cd507",
            },
        },
        "source_artifacts": artifacts,
        "cohort": {
            "species": "Homo sapiens",
            "biological_system": (
                "primary_hepatocytes_isolated_from_histologically_normal_areas_"
                "of_surgical_liver_resections"
            ),
            "donor_count": 7,
            "donors": donors,
            "not_healthy_volunteers": True,
        },
        "measurement_contract": {
            "assay": "quantitative_LC_MS_MS_total_protein_approach_and_proteomic_ruler",
            "protein_entity": "maxquant_protein_group",
            "copy_number_denominator": "per_nucleus",
            "concentration_unit": "pmol_per_mg_total_protein",
            "dna_mass_assumption_pg_per_diploid_nucleus": 6.5,
            "source_zero_or_blank_policy": "nonquantified_null_no_imputation",
            "distinct_groups_may_not_be_collapsed_by_gene": True,
            "copy_number_is_not_surface_abundance": True,
            "copy_number_is_not_active_protein_count": True,
        },
        "source_audit": source_audit,
        "cohort_arithmetic_audit": {
            "donor_weighted_mean_total_protein_pg_per_nucleus": donor_mean_mass,
            "paper_rounded_total_protein_pg_per_reference_cell": 600.0,
            "donor_weighted_mean_sum_of_quantified_group_copies_per_nucleus": donor_mean_copy_sum,
            "paper_rounded_protein_molecules_per_reference_cell": 8_700_000_000.0,
            "article_cell_label_and_supplement_nucleus_denominator_are_not_equivalent_for_binucleate_cells": True,
        },
        "protein_groups": records,
        "integration_gates": {
            "static_donor_abundance_query_ready": True,
            "reference_nucleus_population_initialization_ready": True,
            "donor_specific_cell_initialization_ready": False,
            "binucleate_cell_scaling_ready": False,
            "surface_localized_copy_number_ready": False,
            "transport_active_copy_number_ready": False,
            "protein_turnover_dynamics_ready": False,
            "automatic_flux_coupling": False,
            "literal_molecule_rendering_permitted": False,
            "predictive_ready": False,
        },
        "limitations": [
            "Donors A-G were surgical patients, not healthy volunteers; diagnoses, ages and sex labels remain attached to every observation.",
            "The proteomic ruler reports copies per nucleus. These values are not copies per hepatocyte when a hepatocyte is binucleate.",
            "A MaxQuant protein group can contain multiple accessions or genes; groups are never silently split or merged.",
            "Source zeros and blanks mean non-quantified in this curation and are represented as null, never biological zero and never imputed.",
            "Total abundance does not identify plasma-membrane domain, orientation, surface fraction, transport-active fraction or reaction rate.",
            "Static abundance does not identify synthesis, degradation, trafficking or time-dependent proteostasis.",
            "The atlas is an observational reference and does not define one synthetic average donor or automatically alter cell flux.",
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--table-1", type=Path)
    parser.add_argument("--table-2", type=Path)
    parser.add_argument("--cache-dir", type=Path, default=DEFAULT_CACHE)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    table_1 = args.table_1 or _fetch("supplemental_table_1", args.cache_dir)
    table_2 = args.table_2 or _fetch("supplemental_table_2", args.cache_dir)
    payload = curate(table_1, table_2)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(
        json.dumps(payload, indent=2, ensure_ascii=True) + "\n", encoding="utf-8"
    )
    print(
        f"wrote {args.out} | {payload['source_audit']['quantified_target_rows']} "
        "quantified target protein groups | 7 donors | no imputation"
    )


if __name__ == "__main__":
    main()
